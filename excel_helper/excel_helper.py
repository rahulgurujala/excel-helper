import contextlib
from typing import Any, Dict, List, Literal, Optional, Tuple, TypedDict, Union

import openpyxl
import openpyxl.utils
import openpyxl.utils.exceptions
import pandas as pd
from jinja2 import Template
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet


class ColorScaleKwargs(TypedDict, total=False):
    start_color: str
    end_color: str
    start_type: Literal["min", "max", "percentile", "num", "formula"]
    end_type: Literal["min", "max", "percentile", "num", "formula"]


class ExcelHelper:
    def __init__(self, filename: str):
        self.filename = filename
        self.workbook = None
        self.active_sheet = None

    def open_workbook(self):
        """Open the Excel workbook."""
        self.workbook = openpyxl.load_workbook(self.filename)
        self.active_sheet = self.workbook.active

    def save_workbook(self):
        """Save the Excel workbook."""
        self.workbook.save(self.filename)

    def create_new_workbook(self):
        """Create a new Excel workbook."""
        self.workbook = openpyxl.Workbook()
        self.active_sheet = self.workbook.active

    def select_sheet(self, sheet_name: str):
        """Select a sheet by name."""
        if sheet_name in self.workbook.sheetnames:
            self.active_sheet = self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    def write_cell(self, row: int, col: int, value: Any):
        """Write a value to a specific cell."""
        self.active_sheet.cell(row=row, column=col, value=value)

    def read_cell(self, row: int, col: int) -> Any:
        """Read the value from a specific cell."""
        return self.active_sheet.cell(row=row, column=col).value

    def write_row(self, row: int, data: List[Any]):
        """Write a list of values to a row."""
        for col, value in enumerate(data, start=1):
            self.write_cell(row, col, value)

    def read_row(self, row: int) -> List[Any]:
        """Read all values from a row."""
        return [cell.value for cell in self.active_sheet[row]]

    def write_column(self, col: int, data: List[Any]):
        """Write a list of values to a column."""
        for row, value in enumerate(data, start=1):
            self.write_cell(row, col, value)

    def read_column(self, col: int) -> List[Any]:
        """Read all values from a column."""
        return [cell.value for cell in self.active_sheet[get_column_letter(col)]]

    def write_range(self, start_row: int, start_col: int, data: List[List[Any]]):
        """Write a 2D list of values to a range of cells."""
        for row_offset, row_data in enumerate(data):
            for col_offset, value in enumerate(row_data):
                self.write_cell(start_row + row_offset, start_col + col_offset, value)

    def read_range(
        self, start_row: int, start_col: int, end_row: int, end_col: int
    ) -> List[List[Any]]:
        """Read a range of cells and return a 2D list of values."""
        return [
            [cell.value for cell in row]
            for row in self.active_sheet.iter_rows(
                min_row=start_row, min_col=start_col, max_row=end_row, max_col=end_col
            )
        ]

    def apply_style(self, row: int, col: int, style: Dict[str, Any]):
        """Apply a style to a specific cell."""
        cell = self.active_sheet.cell(row=row, column=col)
        for key, value in style.items():
            setattr(cell, key, value)

    def auto_fit_columns(self):
        """Auto-fit all columns in the active sheet."""
        for column in self.active_sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                with contextlib.suppress(Exception):
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
            adjusted_width = max_length + 2
            self.active_sheet.column_dimensions[column_letter].width = adjusted_width

    def set_formula(self, row: int, col: int, formula: str):
        """Set a formula in a specific cell."""
        self.active_sheet.cell(row=row, column=col, value=formula)

    def get_formula(self, row: int, col: int) -> str:
        """Get the formula from a specific cell."""
        return (
            self.active_sheet.cell(row=row, column=col).data_type == "f"
            and self.active_sheet.cell(row=row, column=col).value
        )

    def copy_formula(self, from_row: int, from_col: int, to_row: int, to_col: int):
        """Copy a formula from one cell to another, adjusting cell references."""
        source_cell = self.active_sheet.cell(row=from_row, column=from_col)
        target_cell = self.active_sheet.cell(row=to_row, column=to_col)

        if source_cell.data_type == "f":
            translated_formula = Translator(
                source_cell.value, origin=source_cell.coordinate
            ).translate_formula(target_cell.coordinate)
            target_cell.value = translated_formula

    def sum_range(
        self,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        result_row: int,
        result_col: int,
    ):
        """Sum a range of cells and put the result in another cell."""
        start_cell = get_column_letter(start_col) + str(start_row)
        end_cell = get_column_letter(end_col) + str(end_row)
        formula = f"=SUM({start_cell}:{end_cell})"
        self.set_formula(result_row, result_col, formula)

    def average_range(
        self,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        result_row: int,
        result_col: int,
    ):
        """Calculate the average of a range of cells and put the result in another cell."""
        start_cell = get_column_letter(start_col) + str(start_row)
        end_cell = get_column_letter(end_col) + str(end_row)
        formula = f"=AVERAGE({start_cell}:{end_cell})"
        self.set_formula(result_row, result_col, formula)

    def count_range(
        self,
        start_row: int,
        start_col: int,
        end_row: int,
        end_col: int,
        result_row: int,
        result_col: int,
    ):
        """Count non-empty cells in a range and put the result in another cell."""
        start_cell = get_column_letter(start_col) + str(start_row)
        end_cell = get_column_letter(end_col) + str(end_row)
        formula = f"=COUNT({start_cell}:{end_cell})"
        self.set_formula(result_row, result_col, formula)

    def if_formula(
        self,
        condition_row: int,
        condition_col: int,
        true_value: Any,
        false_value: Any,
        result_row: int,
        result_col: int,
    ):
        """Set an IF formula in a specific cell."""
        condition_cell = get_column_letter(condition_col) + str(condition_row)
        formula = f'=IF({condition_cell}, "{true_value}", "{false_value}")'
        self.set_formula(result_row, result_col, formula)

    def vlookup(
        self,
        lookup_value_row: int,
        lookup_value_col: int,
        table_start_row: int,
        table_start_col: int,
        table_end_row: int,
        table_end_col: int,
        col_index: int,
        result_row: int,
        result_col: int,
    ):
        """Set a VLOOKUP formula in a specific cell."""
        lookup_value = get_column_letter(lookup_value_col) + str(lookup_value_row)
        table_range = f"{get_column_letter(table_start_col)}{table_start_row}:{get_column_letter(table_end_col)}{table_end_row}"
        formula = f"=VLOOKUP({lookup_value}, {table_range}, {col_index}, FALSE)"
        self.set_formula(result_row, result_col, formula)

    def create_chart(
        self,
        chart_type: str,
        data_range: Union[List[int], Tuple[int, int, int, int]],
        title: str,
        x_axis: str,
        y_axis: str,
        location: str,
    ) -> None:
        """
        Create a chart in the Excel workbook.
        """
        chart = BarChart() if chart_type.lower() == "bar" else None
        # TODO Add more chart types as needed

        if chart:
            data = Reference(
                self.active_sheet,
                min_col=data_range[0],
                min_row=data_range[1],
                max_col=data_range[2],
                max_row=data_range[3],
            )
            chart.add_data(data, titles_from_data=True)
            chart.title = title
            chart.x_axis.title = x_axis
            chart.y_axis.title = y_axis
            self.active_sheet.add_chart(chart, location)

    def create_pivot_table(
        self,
        source_data: List[List[Any]],
        pivot_location: str,
        rows: List[str],
        columns: List[str],
        values: List[str],
    ) -> None:
        """
        Create a pivot table in the Excel workbook.
        """
        pivot_sheet: Worksheet = self.workbook.create_sheet("PivotTable")
        pivot_sheet.cell(row=1, column=1, value="Pivot Table")

        # Create a PivotTable
        pivot_table: Table = Table(
            displayName="PivotTable",
            ref=f"A1:{get_column_letter(len(source_data[0]))}{len(source_data)}",
        )
        pivot_sheet.add_table(pivot_table)

        for row in source_data:
            pivot_sheet.append(row)

        pivot_fields: List[dict] = [
            {
                "sourceField": field,
                "orientation": (
                    "row"
                    if field in rows
                    else "column" if field in columns else "value"
                ),
            }
            for field in rows + columns + values
        ]

        pivot_sheet.pivot_tables.add(
            "PivotTable1",
            f"A1:{get_column_letter(len(source_data[0]))}{len(source_data)}",
            pivot_location,
            pivot_fields,
        )

    def add_data_validation(
        self,
        cell_range: str,
        validation_type: str,
        validation_criteria: str,
        validation_value: Union[str, int, float],
    ) -> None:
        """
        Add data validation to a range of cells.
        """
        dv = DataValidation(
            type=validation_type,
            operator=validation_criteria,
            formula1=validation_value,
        )
        self.active_sheet.add_data_validation(dv)
        dv.add(cell_range)

    def apply_conditional_formatting(
        self,
        cell_range: str,
        rule_type: Literal["color_scale"],
        **kwargs: ColorScaleKwargs,
    ) -> None:
        """
        Apply conditional formatting to a range of cells.
        """
        if rule_type != "color_scale":
            raise ValueError(f"Unsupported rule_type: {rule_type}")
        rule = ColorScaleRule(
            start_color=kwargs.get("start_color", "FFFFFF"),
            end_color=kwargs.get("end_color", "FF0000"),
            start_type=kwargs.get("start_type", "min"),
            end_type=kwargs.get("end_type", "max"),
        )
        self.active_sheet.conditional_formatting.add(cell_range, rule)

    def create_macro(self, macro_name: str, macro_code: str) -> None:
        """
        Create a new macro in the Excel workbook.
        """
        if not hasattr(self.workbook, "vba_archive"):
            raise AttributeError(
                "This workbook doesn't support VBA macros. "
                "Make sure you've created it with keep_vba=True."
            )

        if not self.workbook.vba_archive:
            self.workbook.create_vba_module()

        module = self.workbook.vba_archive.get_or_create_module("Module1")
        module.write(f"Sub {macro_name}()\n{macro_code}\nEnd Sub")

    def run_macro(self, macro_name: str) -> None:
        """
        Run a macro in the Excel workbook.
        """
        if not self._is_windows():
            raise OSError("This method can only be run on Windows.")

        try:
            import win32com.client

            excel: Any = win32com.client.Dispatch("Excel.Application")
            wb: Any = excel.Workbooks.Open(self.filename)
            excel.Application.Run(macro_name)
            wb.Save()
        except Exception as e:
            # sourcery skip: raise-specific-error
            raise Exception(f"Error running macro: {str(e)}") from e
        finally:
            if "excel" in locals():
                excel.Application.Quit()

    def _is_windows(self) -> bool:
        """Check if the current operating system is Windows."""
        import platform

        return platform.system().lower() == "windows"

    def to_dataframe(
        self,
        sheet_name: Union[str, None] = None,
        start_row: int = 1,
        start_col: int = 1,
    ):
        """Convert Excel data to a Pandas DataFrame."""
        if sheet_name:
            self.select_sheet(sheet_name)

        data = self.read_range(
            start_row,
            start_col,
            self.active_sheet.max_row,
            self.active_sheet.max_column,
        )
        return pd.DataFrame(data[1:], columns=data[0])

    def from_dataframe(
        self,
        df: pd.DataFrame,
        sheet_name: Optional[str] = None,
        start_row: int = 1,
        start_col: int = 1,
    ) -> None:
        """
        Write a Pandas DataFrame to the Excel workbook.
        """
        if df.empty:
            raise ValueError("Cannot write an empty DataFrame to Excel.")

        if sheet_name:
            self.select_sheet(sheet_name)

        self.write_range(
            start_row, start_col, [df.columns.tolist()] + df.values.tolist()
        )

    def use_template(
        self, template_file: str, output_file: str, context: Dict[str, Any]
    ) -> None:
        """
        Use an Excel template to generate a report with dynamic data.

        This method loads an Excel template, replaces placeholders with dynamic data,
        and saves the result to a new file.

        Args:
            template_file (str): The path to the Excel template file.
            output_file (str): The path where the generated report will be saved.
            context (Dict[str, Any]): A dictionary containing the data to be inserted into the template.
                Keys should match the placeholders in the template.

        Returns:
            None

        Raises:
            FileNotFoundError: If the template file doesn't exist.
            PermissionError: If there's no write permission for the output file.
            ValueError: If the template file is not a valid Excel file.
        """
        try:
            wb: Workbook = openpyxl.load_workbook(template_file)
            ws: Worksheet = wb.active

            for cell in ws._cells.values():
                if cell.data_type == "s" and "{{" in cell.value and "}}" in cell.value:
                    template = Template(cell.value)
                    cell.value = template.render(context)

            wb.save(output_file)
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Template file not found: {template_file}") from e
        except PermissionError as e:
            raise PermissionError(
                f"Permission denied when trying to save to: {output_file}"
            ) from e
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ValueError(f"Invalid Excel file: {template_file}") from e
